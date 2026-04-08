#!/usr/bin/env python3
"""
将 docs/bazaar.xlsx 按工作表导出为多个 CSV（每表一个 CSV），用于“添加物品表格”导入。

约定：
- 固定 9 列：中文名、英文名、英雄、版本、minTier、Size、CD、标签、效果
- CD 若被 Excel 误存为日期：按 “月/日/年后两位” 还原（例如 2005-07-06 → 7/6/5；若仅月日则 → 7/6）
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.numbers import is_date_format
except ImportError:
    raise SystemExit("请先安装依赖：python -m pip install openpyxl")


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = REPO_ROOT / "docs" / "bazaar.xlsx"
DEFAULT_OUT_DIR = REPO_ROOT / "docs" / "item_sheets_csv"

EXPECTED_HEADER = ["中文名", "英文名", "英雄", "版本", "minTier", "Size", "CD", "标签", "效果"]

LAYOUT_V2_HEADER = ["序号", "物品名", "英文名", "英雄", "历史", "级", "体", "CD", "TAG"]


@dataclass(frozen=True)
class SheetLayout:
    # 0-based indices in the input row to produce the 9 output columns.
    out_map: list[int]
    # 0-based input column index of CD (for date coercion)
    cd_col: int
    # whether the first non-empty row is a header row
    header_is_first_non_empty: bool


def detect_layout(first9_texts: list[str], first10_texts: list[str]) -> SheetLayout:
    # Layout v2: 10 columns, first row like:
    # 序号, 物品名, 英文名, 英雄, 历史, 级, 体, CD, TAG, （第10列为效果/备注）
    # We export 9 columns by mapping:
    # 中文名=物品名, 英文名, 英雄, 版本=历史, minTier=级, Size=体, CD=CD, 标签=TAG, 效果=第10列
    if first9_texts[:9] == LAYOUT_V2_HEADER:
        return SheetLayout(
            out_map=[1, 2, 3, 4, 5, 6, 7, 8, 9],
            cd_col=7,
            header_is_first_non_empty=True,
        )

    # Layout v1: already matches 9-column item table header
    if matches_header(first9_texts):
        return SheetLayout(out_map=list(range(0, 9)), cd_col=6, header_is_first_non_empty=True)

    # Fallback: assume first 9 columns are already the 9 output columns
    return SheetLayout(out_map=list(range(0, 9)), cd_col=6, header_is_first_non_empty=False)


@dataclass(frozen=True)
class ExportStats:
    sheet: str
    out_path: Path
    rows_written: int
    header_skipped: bool


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="将 bazaar.xlsx 按工作表导出为多个 CSV（物品表 9 列格式）。")
    p.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="输入 xlsx 路径（默认 docs/bazaar.xlsx）。")
    p.add_argument(
        "--out-dir",
        default=str(DEFAULT_OUT_DIR),
        help="输出目录（默认 docs/item_sheets_csv/）。",
    )
    p.add_argument("--sheets", default=None, help="仅导出指定工作表（逗号分隔，默认全部）。")
    p.add_argument(
        "--fill-hero-from-sheet",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="当“英雄”列为空时，用工作表名填充（默认开启）。",
    )
    p.add_argument(
        "--inspect",
        action="store_true",
        help="仅打印每个工作表的结构与 CD 样本，不输出 CSV。",
    )
    p.add_argument(
        "--check",
        action="store_true",
        help="导出后做最小校验（列数/关键列非空比例），仅报警不阻断。",
    )
    return p.parse_args()


def sanitize_filename(name: str) -> str:
    # Windows 文件名非法字符：<>:"/\\|?*
    name = re.sub(r'[<>:"/\\\\|?*]+', "_", name).strip()
    name = re.sub(r"\s+", " ", name)
    return name or "sheet"


def is_empty_cell_value(v: object) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def normalize_number(v: float | int) -> str:
    # 版本号等经常被存成 12.0
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def looks_like_excel_anchor_year(y: int) -> bool:
    # Excel 把“仅月日”的显示也可能落到一个锚定年；这里用保守白名单。
    return y in {1899, 1900, 1970, 2000}


def cd_from_date(d: date) -> str:
    # 还原规则：优先输出 m/d/yy；若年份像锚定年则仅输出 m/d。
    m = int(d.month)
    dd = int(d.day)
    y = int(d.year)
    if looks_like_excel_anchor_year(y):
        return f"{m}/{dd}"
    return f"{m}/{dd}/{y % 100}"


def cell_to_text(cell: Cell, *, is_cd: bool) -> str:
    v = cell.value
    if is_empty_cell_value(v):
        return ""

    # CD 列：优先按日期纠偏
    if is_cd:
        if isinstance(v, (datetime, date)):
            return cd_from_date(v.date() if isinstance(v, datetime) else v)
        if is_date_format(cell.number_format):
            # 某些情况下 value 不是 date，但 number_format 标成日期；尽量按显示值的 date 语义处理
            if isinstance(v, (int, float)):
                # Excel serial date；openpyxl 通常会转为 datetime，但这里兜底：直接转字符串以免误判
                return normalize_number(v)
            return str(v).strip()
        if isinstance(v, (int, float)):
            return normalize_number(v)
        return str(v).strip()

    # 非 CD：稳定字符串化
    if isinstance(v, (int, float)):
        return normalize_number(v)
    return str(v).strip()


def row_is_all_empty(cells: list[Cell]) -> bool:
    return all(is_empty_cell_value(c.value) for c in cells)


def matches_header(row_texts: list[str]) -> bool:
    # “高度匹配”：前 9 列里包含这些关键字
    if len(row_texts) < 9:
        return False
    a = [t.strip() for t in row_texts[:9]]
    # 允许 “最低等级”“冷却”等变体：只要关键列对齐即可
    required = {"中文名", "英文名"}
    if not required.issubset(set(a)):
        return False
    # 若第 1 列是 中文名、第 2 列是 英文名，则认为表头
    return a[0] == "中文名" and a[1] == "英文名"


def iter_sheet_rows(ws) -> Iterable[list[Cell]]:
    # 按 max_column 读取，避免行尾缺列导致迭代长度不一致
    max_col = max(9, int(ws.max_column or 9))
    for row in ws.iter_rows(min_row=1, max_row=int(ws.max_row or 1), min_col=1, max_col=max_col):
        yield list(row)


def export_one_sheet(
    ws,
    *,
    sheet_name: str,
    out_dir: Path,
    fill_hero_from_sheet: bool,
    inspect_only: bool,
    check: bool,
) -> Optional[ExportStats]:
    rows = list(iter_sheet_rows(ws))
    if not rows:
        return None

    # 找到首个非空行，用于判断表头与结构
    first_non_empty_idx: Optional[int] = None
    for i, r in enumerate(rows):
        if not row_is_all_empty(r):
            first_non_empty_idx = i
            break

    if first_non_empty_idx is None:
        return None

    first_row = rows[first_non_empty_idx]
    first10 = first_row[:10] if len(first_row) >= 10 else first_row
    first10_texts = [cell_to_text(c, is_cd=False) for c in first10]
    first9_texts = first10_texts[:9] + ([""] * (9 - len(first10_texts[:9])))
    layout = detect_layout(first9_texts, first10_texts)
    header_skipped = layout.header_is_first_non_empty

    # inspect：抓 CD 列样本（第 7 列）
    if inspect_only:
        cd_samples: list[tuple[int, str, str, str, bool]] = []
        for idx in range(
            first_non_empty_idx + (1 if header_skipped else 0),
            min(first_non_empty_idx + 1 + 400, len(rows)),
        ):
            row = rows[idx]
            if len(row) <= layout.cd_col:
                continue
            c = row[layout.cd_col]
            v = c.value
            if is_empty_cell_value(v):
                continue
            is_date = isinstance(v, (datetime, date)) or is_date_format(c.number_format)
            cd_samples.append((idx + 1, type(v).__name__, str(v), str(c.number_format), bool(is_date)))
            if len(cd_samples) >= 6:
                break
        print(
            f"[inspect] sheet={sheet_name} max_row={ws.max_row} max_col={ws.max_column} "
            f"header_skipped={header_skipped} layout_out_map={layout.out_map} first10={first10_texts[:10]} cd_samples={cd_samples}"
        )
        return None

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{sanitize_filename(sheet_name)}.csv"

    rows_written = 0
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)

        for idx, row in enumerate(rows):
            if idx == first_non_empty_idx and header_skipped:
                continue
            if idx < first_non_empty_idx:
                # 跳过首个非空行之前的空白/备注行
                if row_is_all_empty(row):
                    continue
            if row_is_all_empty(row):
                continue

            # 先根据 layout 抽取/映射为 9 列（CD 用 layout.cd_col 做日期纠偏）
            texts: list[str] = []
            for j in layout.out_map:
                if j < len(row):
                    cell = row[j]
                    texts.append(cell_to_text(cell, is_cd=(j == layout.cd_col)))
                else:
                    texts.append("")

            # 对齐到 9 列（layout 已输出 9 列，这里仅兜底）
            if len(texts) < 9:
                texts.extend([""] * (9 - len(texts)))
            if len(texts) > 9:
                texts = texts[:9]

            # 英雄列为空则用工作表名填充
            if fill_hero_from_sheet and texts[2].strip() == "":
                texts[2] = sheet_name

            # 过滤明显的非数据行：中文名/英文名都为空时不导出
            if texts[0].strip() == "" and texts[1].strip() == "":
                continue

            w.writerow(texts[:9])
            rows_written += 1

    if check:
        warn_basic_csv(out_path)

    return ExportStats(sheet=sheet_name, out_path=out_path, rows_written=rows_written, header_skipped=header_skipped)


def warn_basic_csv(csv_path: Path) -> None:
    total = 0
    good_cols = 0
    non_empty_zh_en = 0
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        r = csv.reader(f)
        for row in r:
            total += 1
            if len(row) == 9:
                good_cols += 1
            if len(row) >= 2 and row[0].strip() and row[1].strip():
                non_empty_zh_en += 1
    if total == 0:
        print(f"[check][warn] 空文件：{csv_path}")
        return
    if good_cols != total:
        print(f"[check][warn] {csv_path.name} 存在列数!=9 的行：{total-good_cols}/{total}")
    ratio = non_empty_zh_en / total
    if ratio < 0.9:
        print(f"[check][warn] {csv_path.name} 中文名/英文名均非空比例偏低：{non_empty_zh_en}/{total} ({ratio:.1%})")


def main() -> int:
    args = parse_args()

    xlsx_path = Path(args.xlsx)
    out_dir = Path(args.out_dir)
    if not xlsx_path.is_file():
        print(f"[ERROR] xlsx 不存在：{xlsx_path}")
        return 2

    wb = load_workbook(xlsx_path, data_only=False)
    wanted = None
    if args.sheets:
        wanted = {s.strip() for s in str(args.sheets).split(",") if s.strip()}

    stats: list[ExportStats] = []
    for sheet_name in wb.sheetnames:
        if wanted is not None and sheet_name not in wanted:
            continue
        ws = wb[sheet_name]
        st = export_one_sheet(
            ws,
            sheet_name=sheet_name,
            out_dir=out_dir,
            fill_hero_from_sheet=bool(args.fill_hero_from_sheet),
            inspect_only=bool(args.inspect),
            check=bool(args.check),
        )
        if st is not None:
            stats.append(st)

    if args.inspect:
        return 0

    for st in stats:
        print(f"已导出：{st.sheet} -> {st.out_path}（rows={st.rows_written}, header_skipped={st.header_skipped}）")
    print(f"完成：共导出 {len(stats)} 个工作表到 {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

