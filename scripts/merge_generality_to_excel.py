#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导出 heroes 的物品信息与 generality 扩展列到一个 Excel。

输出包含两个工作表：vanessa / mak。
每表基础列：
- Rank: 从 1 开始的序号（按最高等级扩展列 y 从高到低排序）
- Item: 物品名（Name；烙刀会额外补出加速烙刀/减速烙刀）
- Hero: 英雄名
- Tier: min Tier（B/S/G/D/L）
- Size: S/M/L
- CD: Cooldown（秒；格式同 web 的数字展示）
- TAG: 原生 Tags（不含 derived tags），每行一个，显示中文名
- Desc: 描述（分号替换为换行；占位符按 web pool 逻辑渲染为纯文本阶梯）

扩展列：
- vanessa: l2..l7
- mak: l2..l4
每列来自 out/<hero>/l*/generality.csv 的 generality，经公式变换：
  y = INT(100 * ((x*a)^sqrt(10/b)))
  x: generality.csv 的 generality
  a: 等级补正（l2=5/3, l3=5/4, l4+=1）
  b: generality.csv 的总 item 数量（不含表头）
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
from pathlib import Path
from typing import Any, Iterable


_DURATION_SUFFIX_RE = re.compile(r"^(\d+(?:\.\d+)?)(?:s|_s)$")

_TIER_TO_IDX: dict[str, int] = {"Bronze": 0, "Silver": 1, "Gold": 2, "Diamond": 3, "Legendary": 4}
_TIER_TO_LETTER: dict[str, str] = {"Bronze": "B", "Silver": "S", "Gold": "G", "Diamond": "D", "Legendary": "L"}

_SIZE_TO_LETTER: dict[str, str] = {"Small": "S", "Medium": "M", "Large": "L"}

# 与 app/frontend/src/lib/itemTooltip.ts 一致：原生 tags 英文名 → 中文
_TAG_NAME_ZH: dict[str, str] = {
    "Weapon": "武器",
    "Tool": "工具",
    "Apparel": "服饰",
    "Friend": "伙伴",
    "Food": "食物",
    "Tech": "科技",
    "Property": "地产",
    "Vehicle": "载具",
    "Relic": "遗物",
    "Dragon": "巨龙",
    "Drone": "无人机",
    "Toy": "玩具",
    "Aquatic": "水系",
    "Ray": "射线",
    "Trap": "陷阱",
    "Loot": "战利品",
    "Reagent": "原料",
    "Potion": "药水",
    "Core": "核心",
    "Dinosaur": "恐龙",
}

_MS_DISPLAY_KEYS = frozenset({"Cooldown", "Haste", "Slow", "Freeze", "Charge"})
_DESC_FIELD_ALIASES: dict[str, str] = {
    "ChargeSeconds": "Charge",
    "SlowSeconds": "Slow",
    "FreezeSeconds": "Freeze",
    "HasteSeconds": "Haste",
    "ModifyTargetCount": "ModifyAttributeTargetCount",
}

_YAML_FIELD_ALIASES: dict[str, str] = {
    "cooldown": "Cooldown",
    "cooldownSeconds": "Cooldown",
    "modifytargetcount": "ModifyAttributeTargetCount",
}

_YAML_STRUCT_SPECIAL_KEYS = frozenset(
    {
        "Name",
        "name",
        "Desc",
        "desc",
        "Size",
        "size",
        "Tier",
        "tier",
        "Tags",
        "tags",
        "Abilities",
        "abilities",
        "Auras",
        "auras",
        "overridable",
        "_source_yaml",
        "_hero",
    }
)


def _format_number(n: float) -> str:
    if not math.isfinite(n):
        return ""
    if float(n).is_integer():
        return str(int(n))
    t = f"{n:.4f}".rstrip("0").rstrip(".")
    return t if t else "0"


def _format_seconds_from_ms(ms: float) -> str:
    return _format_number(ms / 1000.0)


def _preprocess_desc(desc: str) -> str:
    # 与 web preprocessDescForTooltip 一致：分号表示换行
    return desc.replace(";", "\n").replace("；", "\n")


def _normalize_placeholder_inner(inner: str) -> tuple[str, bool, bool, bool]:
    s = inner.strip()
    leading_plus = s.startswith("+")
    if leading_plus:
        s = s[1:].strip()
    scale_thousandth = s.endswith(":")
    if scale_thousandth:
        s = s[:-1].strip()
    trailing_percent = s.endswith("%")
    if trailing_percent:
        s = s[:-1]
    key = _DESC_FIELD_ALIASES.get(s, s)
    return key, leading_plus, trailing_percent, scale_thousandth


def _values_all_equal(arr: list[float], eps: float = 1e-6) -> bool:
    if len(arr) <= 1:
        return True
    f = arr[0]
    return all(abs(x - f) < eps for x in arr)


def _slice_tier_values_for_pool(arr: list[float], min_tier: int) -> list[float]:
    mt = max(0, min(4, int(min_tier)))
    if mt >= 4:
        if len(arr) >= 5:
            return [arr[4]]
        return [arr[-1]] if arr else []
    # pool: min_tier..Diamond(3)
    return arr[mt:4]


def _format_one_value_plain(key: str, v: float, scale_thousandth: bool) -> str:
    if scale_thousandth or key in _MS_DISPLAY_KEYS:
        return _format_seconds_from_ms(v)
    return _format_number(v)


def _format_ladder_plain(
    key: str,
    sliced: list[float],
    leading_plus: bool,
    trailing_percent: bool,
    scale_thousandth: bool,
) -> str:
    parts: list[str] = []
    for v in sliced:
        core = _format_one_value_plain(key, v, scale_thousandth)
        s = f"{'+' if leading_plus else ''}{core}{'%' if trailing_percent else ''}"
        parts.append(s)
    return " » ".join(parts)


def _render_desc_plain(desc_template: str, tooltip_attrs: dict[str, list[float]] | None, *, min_tier: int) -> str:
    desc = _preprocess_desc(desc_template)
    re_ph = re.compile(r"\{([^}]+)\}")

    def repl(m: re.Match[str]) -> str:
        inner = m.group(1) or ""
        key, leading_plus, trailing_percent, scale_thousandth = _normalize_placeholder_inner(inner)
        arr = (tooltip_attrs or {}).get(key)
        if not arr:
            return "—"
        sliced = _slice_tier_values_for_pool([float(x) for x in arr], min_tier)
        if not sliced:
            return "—"
        if len(sliced) == 1 or _values_all_equal(sliced):
            core = _format_one_value_plain(key, sliced[0], scale_thousandth)
            return f"{'+' if leading_plus else ''}{core}{'%' if trailing_percent else ''}"
        return _format_ladder_plain(key, sliced, leading_plus, trailing_percent, scale_thousandth)

    return re_ph.sub(repl, desc)


def _tag_label_zh(tag: str) -> str:
    return _TAG_NAME_ZH.get(tag, tag)


def _size_letter(size_raw: str) -> str:
    return _SIZE_TO_LETTER.get(size_raw, size_raw[:1].upper() if size_raw else "")


def _tier_letter(tier_raw: str) -> str:
    return _TIER_TO_LETTER.get(tier_raw, tier_raw[:1].upper() if tier_raw else "")


def _tier_idx(tier_raw: str) -> int:
    return _TIER_TO_IDX.get(tier_raw, 0)


def _read_yaml_items(path: Path) -> tuple[str, list[dict[str, Any]]]:
    try:
        import yaml  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("缺少依赖 PyYAML。请先安装：python -m pip install pyyaml") from e

    with path.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    if not isinstance(data, dict):
        raise ValueError(f"bad yaml root: {path}")
    hero = data.get("hero")
    if not isinstance(hero, str) or not hero.strip():
        raise ValueError(f"missing hero in yaml: {path}")
    items = data.get("items")
    if not isinstance(items, list):
        raise ValueError(f"missing items list in yaml: {path}")
    out: list[dict[str, Any]] = []
    for it in items:
        if isinstance(it, dict):
            out.append(it)
    return hero, out


def _raw_to_storage_number(canon: str, raw: object, *, where: str) -> float:
    if isinstance(raw, bool):
        raise TypeError(f"{where}: 不能为 bool")
    if canon in _MS_DISPLAY_KEYS:
        if isinstance(raw, str):
            s = raw.strip()
            m = _DURATION_SUFFIX_RE.match(s)
            if m:
                return float(m.group(1)) * 1000.0
            raise ValueError(f"{where}: 期望时长字符串（如 3s），实际 {raw!r}")
        if isinstance(raw, (int, float)):
            return float(raw)
        raise TypeError(f"{where}: 不支持的类型 {type(raw)}")
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        m = _DURATION_SUFFIX_RE.match(raw.strip())
        if m:
            return float(m.group(1)) * 1000.0
        raise ValueError(f"{where}: 期望数字或时长字符串，实际 {raw!r}")
    raise TypeError(f"{where}: 不支持的类型 {type(raw)}")


def _iter_tier_values_from_min(vals: list[object], min_tier: int, *, where: str) -> list[tuple[int, object]]:
    if min_tier < 0 or min_tier > 4:
        raise ValueError(f"{where}: minTier 索引非法：{min_tier}")
    if not vals:
        raise ValueError(f"{where}: 升阶列表不能为空")
    out: list[tuple[int, object]] = []
    for i, raw in enumerate(vals):
        tidx = min_tier + i
        if tidx > 4:
            raise ValueError(f"{where}: 升阶项超出 Legendary（minTier={min_tier}，第{i + 1}档对应 tier={tidx}）")
        out.append((tidx, raw))
    last = vals[-1]
    nxt = min_tier + len(vals)
    for tidx in range(nxt, 5):
        out.append((tidx, last))
    return out


def _build_tooltip_attrs(item: dict[str, Any], *, min_tier_idx: int, src: str, item_name: str) -> dict[str, list[float]]:
    """
    生成每个键 → 长度 5 的 tier 数组（0..4）。
    仅为导出 Desc/CD 服务：不要求覆盖全部引擎 ItemKey，只要 YAML 里出现的键可被查到即可。
    """
    out: dict[str, list[float]] = {}
    keys = sorted(k for k in item.keys() if k not in _YAML_STRUCT_SPECIAL_KEYS)
    for yk in keys:
        raw_key = str(yk).strip()
        canon = _YAML_FIELD_ALIASES.get(raw_key, _YAML_FIELD_ALIASES.get(raw_key.lower(), raw_key))
        val = item.get(yk)
        where_f = f"{src} {item_name}.{canon}"
        row = [0.0] * 5
        if isinstance(val, list):
            pairs = _iter_tier_values_from_min(val, min_tier_idx, where=where_f)
            for tidx, raw in pairs:
                row[tidx] = _raw_to_storage_number(canon, raw, where=f"{where_f}[tier={tidx}]")
        else:
            v = _raw_to_storage_number(canon, val, where=where_f)
            for i in range(5):
                row[i] = v
        out[canon] = row
    return out


def _read_generality_csv(path: Path) -> tuple[dict[str, float], int]:
    if not path.is_file():
        raise FileNotFoundError(str(path))
    with path.open("r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        if r.fieldnames is None:
            raise ValueError(f"empty csv: {path}")
        need = {"item", "generality"}
        got = set(r.fieldnames)
        if not need.issubset(got):
            raise ValueError(f"bad header in {path}: expected at least {sorted(need)}, got {r.fieldnames}")
        out: dict[str, float] = {}
        n = 0
        for row in r:
            item = (row.get("item") or "").strip()
            if not item:
                continue
            g_raw = (row.get("generality") or "").strip()
            if g_raw == "":
                continue
            try:
                g = float(g_raw)
            except ValueError as e:
                raise ValueError(f"bad generality value in {path}: item={item!r} generality={g_raw!r}") from e
            out[item] = g
            n += 1
        return out, n


def _level_a(level: str) -> float:
    if level == "l2":
        return 5.0 / 3.0
    if level == "l3":
        return 5.0 / 4.0
    return 1.0


def _generality_y(x: float, *, a: float, b: int) -> int:
    if b <= 0:
        return 0
    if x <= 0:
        return 0
    exp = math.sqrt(10.0 / float(b))
    return int(100.0 * ((x * a) ** exp))


def _format_cd_plain(tooltip_attrs: dict[str, list[float]] | None, *, min_tier: int) -> str:
    arr = (tooltip_attrs or {}).get("Cooldown")
    if not arr:
        return ""
    if not any(float(v) > 0 for v in arr):
        return ""
    sliced = _slice_tier_values_for_pool([float(x) for x in arr], min_tier)
    if not sliced:
        return ""
    if len(sliced) == 1 or _values_all_equal(sliced):
        return _format_seconds_from_ms(sliced[0])
    return " » ".join(_format_seconds_from_ms(v) for v in sliced)


def _join_lines(lines: Iterable[str]) -> str:
    cleaned = [s for s in (x.strip() for x in lines) if s]
    return "\n".join(cleaned)


def _hero_sheet_name(hero_key: str) -> str:
    return hero_key.lower()


def _hero_title(hero_key: str) -> str:
    # Hero 列按需求输出英文英雄名（Vanessa / Mak）
    return hero_key[:1].upper() + hero_key[1:].lower()


def _find_item(items: list[dict[str, Any]], name_zh: str) -> dict[str, Any] | None:
    for it in items:
        if (it.get("Name") or "").strip() == name_zh:
            return it
    return None


def _build_rows_for_hero(
    *,
    hero_key: str,
    hero_title: str,
    yaml_items: list[dict[str, Any]],
    repo_root: Path,
    levels: list[str],
    extra_include_items: list[str],
    base_only_items: set[str],
    clone_from_item: dict[str, str],
) -> tuple[list[str], list[list[Any]]]:
    # 读取各等级 generality
    level_maps: dict[str, dict[str, float]] = {}
    level_counts: dict[str, int] = {}
    for lv in levels:
        p = repo_root / "out" / hero_key / lv / "generality.csv"
        mp, n = _read_generality_csv(p)
        level_maps[lv] = mp
        level_counts[lv] = n

    # 基础行：来自 YAML
    base: dict[str, dict[str, Any]] = {}
    for it in yaml_items:
        name = (it.get("Name") or "").strip()
        if not name:
            continue
        base[name] = it

    # 补齐需要包含的 item（可能不在 YAML 中）
    for name in extra_include_items:
        if name in base:
            continue
        src = clone_from_item.get(name)
        if src and src in base:
            cloned = dict(base[src])
            cloned["Name"] = name
            base[name] = cloned

    # 生成每个 item 的基础列与扩展列（先不排序）
    highest_level = levels[-1] if levels else ""
    header = ["Rank", "Item", "Hero", "Tier", "Size", "CD", "TAG", "Desc"] + levels
    rows: list[tuple[int, str, list[Any]]] = []
    for name, it in base.items():
        tier_raw = str(it.get("Tier") or "").strip()
        size_raw = str(it.get("Size") or "").strip()
        tags = it.get("Tags") or []
        if not isinstance(tags, list):
            tags = []
        tags_zh = _join_lines(_tag_label_zh(str(t)) for t in tags)
        min_tier = _tier_idx(tier_raw)
        tooltip_attrs = _build_tooltip_attrs(it, min_tier_idx=min_tier, src=f"data/items/{hero_key}.yaml", item_name=name)
        cd = _format_cd_plain(tooltip_attrs, min_tier=min_tier)
        desc = _render_desc_plain(str(it.get("Desc") or ""), tooltip_attrs, min_tier=min_tier)

        ext_vals: list[Any] = []
        if name in base_only_items:
            ext_vals = [""] * len(levels)
            top_score = -1
        else:
            top_score = -1
            for lv in levels:
                mp = level_maps.get(lv, {})
                x = mp.get(name)
                if x is None:
                    ext_vals.append("")
                    continue
                y = _generality_y(float(x), a=_level_a(lv), b=level_counts.get(lv, 0))
                ext_vals.append(y)
                if lv == highest_level:
                    top_score = y

        row = [
            None,  # Rank placeholder
            name,
            hero_title,
            _tier_letter(tier_raw),
            _size_letter(size_raw),
            cd,
            tags_zh,
            desc,
            *ext_vals,
        ]
        rows.append((top_score, name, row))

    # 排序：最高等级 y 值 desc
    rows.sort(key=lambda t: (t[0], t[1]), reverse=True)
    out_rows: list[list[Any]] = []
    for i, (_, _, row) in enumerate(rows, start=1):
        row[0] = i
        out_rows.append(row)
    return header, out_rows


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    default_repo = script_dir.parent

    p = argparse.ArgumentParser(description="导出 vanessa/mak 物品与 generality 扩展列到一个 xlsx。")
    p.add_argument(
        "--repo-root",
        type=Path,
        default=default_repo,
        help="仓库根目录（默认：本脚本上级目录）",
    )
    p.add_argument(
        "--out",
        type=Path,
        default=None,
        help="输出 xlsx 路径（默认：<repo-root>/docs/generality_export.xlsx）",
    )
    args = p.parse_args()

    repo_root: Path = args.repo_root.resolve()
    out_path: Path = (args.out or (repo_root / "docs" / "generality_export.xlsx")).resolve()

    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment  # type: ignore
    except Exception:
        print(
            "error: 缺少依赖 openpyxl。请先安装：\n"
            "  python -m pip install openpyxl\n",
            file=sys.stderr,
        )
        return 2

    try:
        vanessa_title, vanessa_items = _read_yaml_items(repo_root / "data" / "items" / "vanessa.yaml")
        mak_title, mak_items = _read_yaml_items(repo_root / "data" / "items" / "mak.yaml")
    except Exception as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    wb = Workbook()
    # 删除默认 sheet，稍后创建两个
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    # 特殊规则
    extra_include_items_vanessa = ["加速烙刀", "减速烙刀"]
    base_only_items_common = {"烙刀", "产药药水", "催化剂"}
    clone_from = {"加速烙刀": "烙刀", "减速烙刀": "烙刀"}

    try:
        vanessa_header, vanessa_rows = _build_rows_for_hero(
            hero_key="vanessa",
            hero_title=vanessa_title,
            yaml_items=vanessa_items,
            repo_root=repo_root,
            levels=["l2", "l3", "l4", "l5", "l6", "l7"],
            extra_include_items=extra_include_items_vanessa,
            base_only_items=base_only_items_common,
            clone_from_item=clone_from,
        )
        mak_header, mak_rows = _build_rows_for_hero(
            hero_key="mak",
            hero_title=mak_title,
            yaml_items=mak_items,
            repo_root=repo_root,
            levels=["l2", "l3", "l4"],
            extra_include_items=[],
            base_only_items=base_only_items_common,
            clone_from_item={},
        )
    except Exception as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    wrap = Alignment(wrap_text=True)

    for sheet_name, header, rows in [
        (_hero_sheet_name("vanessa"), vanessa_header, vanessa_rows),
        (_hero_sheet_name("mak"), mak_header, mak_rows),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(header)
        for r in rows:
            ws.append(r)
        # TAG(7) / Desc(8) 列开启换行（1-based）
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=8):
            for cell in row:
                cell.alignment = wrap

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"wrote {len(vanessa_rows)} vanessa items, {len(mak_rows)} mak items to {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

